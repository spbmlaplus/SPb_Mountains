#!/usr/bin/env python3
"""Export new_legend/longread (2).xlsx sheet LONGREAD to new_legend/LONGREAD.csv.

Only visible rows are exported (hidden rows in Excel are skipped).
Cell values are plain text (rich text flattened).
"""
import csv
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "new_legend" / "longread (2).xlsx"
SHEET_NAME = "LONGREAD"
OUT_PATH = ROOT / "new_legend" / "LONGREAD.csv"


def cell_plain(value) -> str:
    if value is None:
        return ""
    if isinstance(value, CellRichText):
        return "".join(str(part) for part in value)
    return str(value)


def is_row_hidden(ws, row_idx: int) -> bool:
    dim = ws.row_dimensions.get(row_idx)
    return bool(dim and dim.hidden)


def is_col_hidden(ws, col_idx: int) -> bool:
    letter = openpyxl.utils.get_column_letter(col_idx)
    dim = ws.column_dimensions.get(letter)
    return bool(dim and dim.hidden)


def main() -> None:
    if not XLSX_PATH.is_file():
        raise SystemExit(f"Missing xlsx: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=False)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    max_col = ws.max_column or 0
    visible_cols = [c for c in range(1, max_col + 1) if not is_col_hidden(ws, c)]

    rows_out: list[list[str]] = []
    for row_idx in range(1, (ws.max_row or 0) + 1):
        if is_row_hidden(ws, row_idx):
            continue
        rows_out.append([cell_plain(ws.cell(row=row_idx, column=c).value) for c in visible_cols])

    while rows_out and not any(cell.strip() for cell in rows_out[-1]):
        rows_out.pop()

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUT_PATH.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows_out)

    print(f"Wrote {len(rows_out)} rows to {OUT_PATH}")


if __name__ == "__main__":
    main()
