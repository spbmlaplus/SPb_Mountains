#!/usr/bin/env python3
"""Generate src/fallbackLongread.ts from new_legend/longread (2).xlsx.

The xlsx is the design source of truth: unlike the flattened LONGREAD.csv it
preserves per-run bold formatting in the `Description` column. Bold runs are
emitted as <b>...</b> so the longread renders them bold + underlined (App.css).

Faithful transform of the documented rules (read.txt):
  - `Chapter` / `Subtitle` are fill-forwarded: a value persists until the next
    non-empty cell. A new (changed) chapter resets the carried subtitle.
  - One `Description` cell = one paragraph (kept verbatim, incl. newlines; bold
    runs wrapped in <b>).
  - `line`  1 -> divider above; 0 -> dividers above and below.
  - `Media_Link` + `Media Link_type` (left|right) + `name_Media Link` (caption).
  - `fact` -> pink "ФАКТ #N" panel.
  - `Zoom` (name_layer) + `Zoom_view` (1 = target layer stays hidden).
  - `id_layer_base`, `id _map`.

Rows that carry no displayable content (no paragraph, media or fact) are not
emitted, but still advance the chapter/subtitle fill-forward state.
"""
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "new_legend" / "longread (2).xlsx"
SHEET_NAME = "LONGREAD"
OUT_PATH = ROOT / "src" / "fallbackLongread.ts"

# Chapter normalization: the source interleaves "Как устроен амфитеатр" rows
# with "Горный Петербург"; per the design owner they are one chapter — merge
# them so the longread no longer jumps between chapters.
CHAPTER_RENAME = {
    "Как устроен амфитеатр": "Горный Петербург",
}

# Media links removed from the longread (design owner, phase 3).
SKIP_MEDIA_LINKS = {"Finns 2021_legend.png"}

# Facts not yet in xlsx `fact` column (phase 4 hand-tuned; preserve on regen).
# Runtime id_map fixes: xlsx id _map vs LAYERS.csv / PLAN-QA-2026-07.md (see agentic_dev/HANDOFF-2026-07-03.md).
# Each rule: (match in subtitle | media | description) -> id_map.
ID_MAP_OVERRIDES: tuple[tuple[str, int], ...] = (
    ("landscape_7.png", 11),  # geology end — landscape_7 lives in id_map 11, not legacy 12
    ("Pietarin valot", 13),
    ("Призрак Ингерманландии", 14),
    # Vomitorii paragraph — first id_map=7 item (not conclusion longread-18)
    ("долины рек — Тосны, Ижоры, Дудергофки", 7),
)


def _resolve_id_map(
    raw_id_map: int | None,
    subtitle: str,
    media_link: str,
    description_plain: str,
) -> int | None:
    if raw_id_map is None:
        return None
    haystack = f"{subtitle}\n{media_link}\n{description_plain}"
    for needle, mapped in ID_MAP_OVERRIDES:
        if needle in haystack:
            return mapped
    # PLAN-QA: intro / recap — id_map 1 (mount), xlsx still has legacy 2
    if raw_id_map == 2:
        return 1
    return raw_id_map


EXTRA_FACTS_BY_SNIPPET = (
    (
        "прогулки аристократов по паркам",
        "Мы собрали для вас искуссно отображенные горные фасады Петербурга. Давайте знакомится с живописным горным городом!",
    ),
    (
        "сами совершить хайкинг",
        "Исследуйте горы! Проходите маршруты на велосипедах или пешком! Прихватите с собой камеры и следуйте за горными видами Петербурга! А пока мы для вас собрали наши любимые маршруты и фотографии с них.",
    ),
)


def _belveder_allowed(subtitle: str, chapter: str) -> bool:
    combined = f"{subtitle} {chapter}".lower()
    return "николай" in combined or "бельведер" in combined


def _esc(text: str) -> str:
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def cell_plain(value) -> str:
    """Plain text of a cell (rich text flattened, numbers stringified)."""
    if value is None:
        return ""
    if isinstance(value, CellRichText):
        return "".join(str(part) for part in value)
    return str(value)


def cell_html(value) -> str:
    """HTML of a cell: bold runs -> <b>...</b>, everything else escaped."""
    if value is None:
        return ""
    if not isinstance(value, CellRichText):
        return _esc(str(value))

    # Collect (is_bold, text) then coalesce adjacent runs of the same weight so
    # we don't emit fragmented </b><b> seams (e.g. a bold word + bold space).
    runs: list[list] = []
    for part in value:
        is_bold = isinstance(part, TextBlock) and part.font is not None and bool(part.font.b)
        text = str(part)
        if runs and runs[-1][0] == is_bold:
            runs[-1][1] += text
        else:
            runs.append([is_bold, text])

    out = []
    for is_bold, text in runs:
        esc = _esc(text)
        out.append(f"<b>{esc}</b>" if is_bold else esc)
    return "".join(out)


def _int(value):
    text = cell_plain(value).strip()
    if not text:
        return None
    # Numeric cells may arrive as "5" or "5.0"; normalize both.
    if re.fullmatch(r"-?\d+(\.0+)?", text):
        return int(float(text))
    return int(text) if text.lstrip("-").isdigit() else None


def _slug(text: str, index: int) -> str:
    base = re.sub(r"[^a-z0-9a-яё]+", "-", text.lower(), flags=re.IGNORECASE)[:32].strip("-")
    return f"longread-{index}-{base}" if base else f"longread-{index}"


def main() -> None:
    wb = openpyxl.load_workbook(XLSX_PATH, rich_text=True, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    header = [cell_plain(c.value).strip() for c in ws[1]]
    col = {name: idx for idx, name in enumerate(header)}

    def get(row, name):
        idx = col.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx].value

    items: list[dict] = []
    cur_chapter = ""
    cur_subtitle = ""
    cur_id_layer_base = None
    cur_id_map = None

    rows = list(ws.iter_rows(min_row=2))
    for index, row in enumerate(rows):
        raw_chapter = cell_plain(get(row, "Chapter")).strip()
        raw_chapter = CHAPTER_RENAME.get(raw_chapter, raw_chapter)
        raw_subtitle = cell_plain(get(row, "Subtitle")).strip()
        chapter_just_changed = False

        if raw_chapter and raw_chapter != cur_chapter:
            cur_chapter = raw_chapter
            cur_subtitle = ""  # a new chapter drops the previous subtitle
            chapter_just_changed = True
        if raw_subtitle:
            cur_subtitle = raw_subtitle

        desc_value = get(row, "Description")
        description_plain = cell_plain(desc_value).strip()
        paragraphs = [cell_html(desc_value).strip()] if description_plain else []

        media_link = cell_plain(get(row, "Media_Link")).strip()
        media = None
        if media_link:
            skip = media_link in SKIP_MEDIA_LINKS
            if "Belveder" in media_link and not _belveder_allowed(cur_subtitle, cur_chapter):
                skip = True
            if not skip:
                raw_type = cell_plain(get(row, "Media Link_type")).strip().lower()
                side = "right" if raw_type.startswith("right") else "left"
                caption = cell_plain(get(row, "name_Media Link")).strip()
                # Belveder in Nikolay I section: full-width below the first paragraph.
                if "Belveder" in media_link:
                    side = "full"
                media = {"link": media_link, "side": side}
                if caption:
                    media["caption"] = caption

        fact = cell_plain(get(row, "fact")).strip()
        if not fact:
            for snippet, text in EXTRA_FACTS_BY_SNIPPET:
                if snippet in description_plain:
                    fact = text
                    break

        zoom_layer = cell_plain(get(row, "Zoom")).strip()
        zoom = None
        if zoom_layer:
            zoom = {"layer": zoom_layer, "hidden": cell_plain(get(row, "Zoom_view")).strip() == "1"}

        line = _int(get(row, "line"))
        raw_id_layer_base = _int(get(row, "id_layer_base"))
        raw_id_map = _int(get(row, "id _map"))
        if raw_id_layer_base is not None:
            cur_id_layer_base = raw_id_layer_base
        if raw_id_map is not None:
            cur_id_map = _resolve_id_map(
                raw_id_map,
                cur_subtitle,
                media_link,
                description_plain,
            )
        id_layer_base = cur_id_layer_base
        id_map = cur_id_map

        # Pure fill-forward rows (heading only) advance state but aren't shown.
        if not paragraphs and not media and not fact:
            continue

        title = cur_subtitle or cur_chapter or f"Раздел {index + 1}"
        entry: dict = {
            "id": _slug(cur_subtitle or cur_chapter or str(index), index),
            "title": title,
            "description": "\n\n".join(paragraphs),
            "fileList": [],
            "paragraphs": paragraphs,
            "base_id": 1,
        }
        if cur_chapter:
            entry["chapter"] = cur_chapter
        if cur_subtitle:
            entry["subtitle"] = cur_subtitle
        if line is not None:
            entry["line"] = line
        if media:
            entry["media"] = media
            entry["mediaLink"] = media["link"]
        if fact:
            entry["fact"] = fact
        if zoom:
            entry["zoom"] = zoom
        elif chapter_just_changed and cur_chapter == "Высочайшие наблюдатели":
            entry["zoom"] = {"layer": "amphitheater_bound", "hidden": False}
        if id_layer_base is not None:
            entry["id_layer_base"] = id_layer_base
        if id_map is not None:
            entry["id_map"] = id_map
        items.append(entry)

    OUT_PATH.write_text(
        "// AUTO-GENERATED by scripts/gen-fallback-longread.py from new_legend/longread (2).xlsx.\n"
        "// Do not hand-edit; rerun the generator.\n"
        "import type { ContentItem } from './contentTypes'\n\n"
        f"export const fallbackLongreadItems: ContentItem[] = {json.dumps(items, ensure_ascii=False, indent=2)}\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(items)} items to {OUT_PATH}")


if __name__ == "__main__":
    main()
